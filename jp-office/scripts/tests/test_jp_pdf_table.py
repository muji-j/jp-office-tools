"""jp_pdf: PDF 表抽出(extract_tables)+マークダウン変換+csv/xlsx出力のテスト。"""
from pathlib import Path
import jp_pdf


def _make_pdf(path):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", size=12)
    with pdf.table() as table:
        for data_row in [["Name", "Age", "City"], ["Alice", "30", "Tokyo"], ["Bob", "25", "Osaka"]]:
            r = table.row()
            for datum in data_row:
                r.cell(datum)
    pdf.output(str(path))


def test_extract_tables_basic(tmp_path):
    f = tmp_path / "t.pdf"
    _make_pdf(f)
    tables = jp_pdf.extract_tables(str(f))
    assert len(tables) >= 1
    t = tables[0]
    assert t["page"] == 1 and t["index"] == 1
    assert t["rows"][0] == ["Name", "Age", "City"]
    assert ["Bob", "25", "Osaka"] in t["rows"]


def test_extract_tables_none_cells_safe(tmp_path):
    # None セルが混じっても "" に正規化されクラッシュしない
    rows = jp_pdf._normalize_rows([["a", None, "c"], [None, "b"]])
    assert rows == [["a", "", "c"], ["", "b"]]


def test_tables_to_markdown_escapes_pipe():
    tables = [{"page": 2, "index": 1, "rows": [["列|A", "B"], ["x", "y"]]}]
    md = jp_pdf.tables_to_markdown(tables)
    assert "### p.2 表1" in md
    assert "列\\|A" in md
    assert "| --- | --- |" in md


def test_tables_to_markdown_empty():
    assert jp_pdf.tables_to_markdown([]) == "(表が見つかりませんでした)"


def test_tables_to_files_csv(tmp_path):
    tables = [{"page": 1, "index": 1, "rows": [["a", "b"], ["1", "2"]]},
              {"page": 1, "index": 2, "rows": [["x"], ["y"]]}]
    out = tmp_path / "res.csv"
    saved = jp_pdf.tables_to_files(tables, str(out), "csv")
    assert len(saved) == 2
    p1 = Path(saved[0])
    assert p1.exists() and p1.name == "res_p1_t1.csv"
    assert "a,b" in p1.read_text(encoding="utf-8-sig")


def test_tables_to_files_xlsx(tmp_path):
    tables = [{"page": 1, "index": 1, "rows": [["a", "b"], ["1", "2"]]}]
    out = tmp_path / "res.xlsx"
    saved = jp_pdf.tables_to_files(tables, str(out), "xlsx")
    assert len(saved) == 1 and Path(saved[0]).exists()
    import openpyxl
    wb = openpyxl.load_workbook(saved[0])
    assert "p1_t1" in wb.sheetnames


def test_tables_to_files_empty_csv_no_file(tmp_path):
    # 表が無ければファイルを作らず空リストを返す(csv)
    out = tmp_path / "res.csv"
    saved = jp_pdf.tables_to_files([], str(out), "csv")
    assert saved == []
    assert not out.exists()
    assert list(tmp_path.iterdir()) == []


def test_tables_to_files_empty_xlsx_no_file(tmp_path):
    # 表が無ければファイルを作らず空リストを返す(xlsx、内容の無い Sheet1 を生成しない)
    out = tmp_path / "res.xlsx"
    saved = jp_pdf.tables_to_files([], str(out), "xlsx")
    assert saved == []
    assert not out.exists()
    assert list(tmp_path.iterdir()) == []
